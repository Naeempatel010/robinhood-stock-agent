from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------------ #
# Helpers                                                             #
# ------------------------------------------------------------------ #

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="2E4057")
_ALT_FILL = PatternFill("solid", fgColor="F0F4F8")


def _style_header(ws):
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def _alt_row(ws, i):
    if i % 2 == 0:
        for cell in ws[ws.max_row]:
            cell.fill = _ALT_FILL


# ------------------------------------------------------------------ #
# Text report                                                         #
# ------------------------------------------------------------------ #

def write_txt(snapshot, analyses, historical_comparison, path="analysis.txt") -> None:
    ps = snapshot["portfolio_summary"]
    tickers = list(ps["holdings"].keys())
    collected_at = datetime.fromtimestamp(snapshot["timestamp_ms"] / 1000).strftime("%Y-%m-%d %H:%M:%S")

    lines = []
    lines.append("ROBINHOOD PORTFOLIO ANALYSIS REPORT")
    lines.append(f"Generated:     {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"Data snapshot: {collected_at}")
    lines.append("=" * 60)

    if historical_comparison:
        lines.append(f"\n[PORTFOLIO CHANGE SINCE LAST RUN ({historical_comparison['previous_date']})]")
        lines.append(
            f"  Total equity: ${historical_comparison['previous_equity']:,.2f} "
            f"→ ${historical_comparison['current_equity']:,.2f}  "
            f"({historical_comparison['equity_change_pct']:+.2f}%)"
        )
        for t, ch in historical_comparison["holdings_changes"].items():
            lines.append(
                f"  {t}: ${ch['prev_equity']:,.2f} → ${ch['curr_equity']:,.2f}  ({ch['change_pct']:+.2f}%)"
            )

    sections = [
        ("1. FUNDAMENTAL + SHORT INTEREST + INSIDER ANALYSIS", "fundamental"),
        ("2. TECHNICAL + CORRELATION + DIVERSIFICATION ANALYSIS", "technical"),
        ("3. DCF VALUATION ANALYSIS", "dcf"),
        ("4. MARKET SENTIMENT", "sentiment"),
        ("5. OVERALL MARKET OPINION", "market_opinion"),
        ("6. HOT STOCKS ANALYSIS", "hot_stocks"),
    ]
    for heading, key in sections:
        content = analyses.get(key, "")
        lines.append(f"\n\n[{heading}]")
        lines.append(content if content else "— Skipped (budget limit) —")

    lines.append("\n\n" + "=" * 60)
    lines.append("[FINAL SUMMARY & RECOMMENDATIONS]")
    lines.append("=" * 60)
    summary = analyses.get("final_summary", "")
    lines.append(summary if summary else "— Final summary not generated (budget limit) —")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


# ------------------------------------------------------------------ #
# Excel workbook                                                      #
# ------------------------------------------------------------------ #

def write_xlsx(snapshot, analyses, history, path="portfolio_analysis.xlsx") -> None:
    ps = snapshot["portfolio_summary"]
    tickers = list(ps["holdings"].keys())
    technical_data = snapshot["technical_data"]
    fundamental_data = snapshot["fundamental_data"]
    dcf_results = snapshot["dcf_results"]
    short_interest_data = snapshot["short_interest_data"]
    insider_data = snapshot["insider_data"]
    correlation_data = snapshot["correlation_data"]

    wb = Workbook()

    # --- Sheet 1: Holdings ---
    ws1 = wb.active
    ws1.title = "Holdings"
    ws1.append(["Ticker", "Shares", "Avg Buy Price", "Current Price", "Equity", "Gain/Loss $", "Gain/Loss %"])
    _style_header(ws1)
    for i, (t, h) in enumerate(ps["holdings"].items()):
        ws1.append([
            t,
            h["shares"],
            h["avg_buy_price"],
            technical_data.get(t, {}).get("current_price", ""),
            h["equity"],
            h["equity_change"],
            h["percent_change"],
        ])
        _alt_row(ws1, i)
    _auto_width(ws1)

    # --- Sheet 2: Fundamental ---
    ws2 = wb.create_sheet("Fundamental")
    ws2.append(["Ticker", "Name", "Sector", "P/E", "Fwd P/E", "P/B", "EPS", "Rev Growth",
                "Profit Margin", "ROE", "D/E", "Beta", "Analyst Rating", "Target Price"])
    _style_header(ws2)
    for i, (t, d) in enumerate(fundamental_data.items()):
        ws2.append([
            t, d.get("name"), d.get("sector"), d.get("pe_ratio"), d.get("forward_pe"),
            d.get("pb_ratio"), d.get("eps"), d.get("revenue_growth"), d.get("profit_margin"),
            d.get("return_on_equity"), d.get("debt_to_equity"), d.get("beta"),
            d.get("analyst_rating"), d.get("target_price"),
        ])
        _alt_row(ws2, i)
    _auto_width(ws2)

    # --- Sheet 3: Technical ---
    ws3 = wb.create_sheet("Technical")
    ws3.append(["Ticker", "Price", "RSI", "MACD", "MACD Signal", "BB Upper", "BB Lower",
                "SMA 50", "SMA 200", "1M %", "3M %", "6M %"])
    _style_header(ws3)
    for i, (t, d) in enumerate(technical_data.items()):
        ws3.append([
            t, d.get("current_price"), d.get("rsi_14"), d.get("macd"), d.get("macd_signal"),
            d.get("bb_upper"), d.get("bb_lower"), d.get("sma_50"), d.get("sma_200"),
            d.get("change_1m_pct"), d.get("change_3m_pct"), d.get("change_6m_pct"),
        ])
        _alt_row(ws3, i)
    _auto_width(ws3)

    # --- Sheet 4: DCF Valuation ---
    ws4 = wb.create_sheet("DCF Valuation")
    ws4.append(["Ticker", "Intrinsic Value", "Current Price", "Margin of Safety %", "Verdict"])
    _style_header(ws4)
    for i, (t, d) in enumerate(dcf_results.items()):
        if "note" in d:
            ws4.append([t, "N/A", "N/A", "N/A", d["note"]])
        else:
            ws4.append([t, d.get("intrinsic_value"), d.get("current_price"),
                        d.get("margin_of_safety_pct"), d.get("verdict")])
        _alt_row(ws4, i)
    _auto_width(ws4)

    # --- Sheet 5: Short Interest ---
    ws5 = wb.create_sheet("Short Interest")
    ws5.append(["Ticker", "Name", "Shares Short", "Days to Cover", "Short % of Float",
                "Prior Month Shares Short"])
    _style_header(ws5)
    for i, (t, d) in enumerate(short_interest_data.items()):
        ws5.append([
            t, d.get("name"), d.get("shares_short"), d.get("short_ratio_days"),
            d.get("short_percent_of_float"), d.get("shares_short_prior_month"),
        ])
        _alt_row(ws5, i)
    _auto_width(ws5)

    # --- Sheet 6: Insider Trading ---
    ws6 = wb.create_sheet("Insider Trading")
    ws6.append(["Ticker", "Transaction Details"])
    _style_header(ws6)
    for t, records in insider_data.items():
        if records:
            for rec in records:
                ws6.append([t, str(rec)])
        else:
            ws6.append([t, "No recent insider transactions found"])
    _auto_width(ws6)

    # --- Sheet 7: Correlation ---
    ws7 = wb.create_sheet("Correlation")
    if correlation_data.get("matrix"):
        corr_tickers = list(correlation_data["matrix"].keys())
        ws7.append([""] + corr_tickers)
        _style_header(ws7)
        for t in corr_tickers:
            ws7.append([t] + [round(correlation_data["matrix"][t].get(t2, 0), 3) for t2 in corr_tickers])
        ws7.append([])
        ws7.append(["Avg Pairwise Correlation", correlation_data.get("avg_correlation")])
        ws7.append(["Diversification Score (0-10)", correlation_data.get("diversification_score")])
    _auto_width(ws7)

    # --- Sheet 8: History ---
    ws8 = wb.create_sheet("History")
    ws8.append(["Date", "Total Equity", "Cash"] + tickers)
    _style_header(ws8)
    for run in history:
        row = [run["date"][:10], run["equity"], run["cash"]]
        for t in tickers:
            row.append(run["holdings"].get(t, {}).get("equity", ""))
        ws8.append(row)
    _auto_width(ws8)

    wb.save(path)
